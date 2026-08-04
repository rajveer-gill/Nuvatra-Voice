"""Read a service menu out of a spreadsheet.

Salons on another booking system can export their service list to Excel — Zenoti
can export everything except appointments — so this turns that file into our service
rows instead of making someone retype fifty of them by hand. At thirty-odd locations
that difference decides whether onboarding is feasible.

Deliberately forgiving about layout: the header row is found by looking for a
name-ish column rather than assuming row 1, and column names are matched loosely,
because every system labels these differently ("ServiceName", "Service Name",
"Service"). Anything it can't read is reported rather than guessed at — the caller
shows the rows for confirmation before they're saved.
"""

from __future__ import annotations

import io
import logging
import os
import re
from typing import Any, Optional

logger = logging.getLogger("nuvatra")

# Same lever as the appointment import — one env var to roll back or A/B.
IMPORT_LLM_MODEL = (
    os.getenv("IMPORT_LLM_MODEL") or os.getenv("VOICE_LLM_MODEL") or "gpt-4o-mini"
).strip()

MAX_ROWS = 500
MAX_BYTES = 5 * 1024 * 1024  # a service list is kilobytes; anything larger is a mistake

# Column header aliases, lowercased and stripped of non-letters for matching.
# Deliberately NOT "description" — a real export often has both a name column and a
# "Service Description" column, and matching the latter picks prose instead of names.
_NAME_KEYS = {"servicename", "name", "service", "servicerequested"}
_PRICE_KEYS = {"price", "cost", "amount", "rate", "listprice"}
_DURATION_KEYS = {"servicetime", "servicetimeinminutes", "duration", "durationminutes",
                  "time", "minutes", "length"}
_CATEGORY_KEYS = {"category", "subcategory", "servicecategory", "group", "type"}
_CODE_KEYS = {"servicecode", "code", "sku", "id"}


def _key(s: Any) -> str:
    return re.sub(r"[^a-z]", "", str(s or "").lower())


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return max(0.0, float(v))
    m = re.search(r"(\d+(?:\.\d+)?)", str(v).replace(",", ""))
    return max(0.0, float(m.group(1))) if m else 0.0


def _to_int(v: Any, default: int) -> int:
    if isinstance(v, (int, float)):
        n = int(v)
    else:
        m = re.search(r"(\d+)", str(v or ""))
        n = int(m.group(1)) if m else default
    return n


def _looks_like_addon(name: str, duration: int, category: str) -> bool:
    """Best-effort guess, shown as a pre-ticked box the user can correct.

    Add-ons betray themselves two ways: the name says so ("Hot tool Add-on",
    "Master Stylist 5", the length charges), or the row has no duration because it's
    a charge rather than a booking. Never authoritative — the preview is editable
    precisely because this is a heuristic.
    """
    n = name.lower()
    if any(w in n for w in ("add-on", "add on", "addon", "master stylist")):
        return True
    if any(w in n for w in ("length", "density")) and duration == 0:
        return True
    if "conditioner" in n and duration <= 10:
        return True
    if duration == 0 and any(w in n for w in ("additional", "charge", "upgrade")):
        return True
    return False


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    """Locate the header row and map our fields to column indexes.

    Scans the first 20 rows rather than assuming row 1 — exports routinely carry a
    title or blank rows above the real header. Returns (-1, {}) when no row has
    something name-like in it.
    """
    for i, row in enumerate(rows[:20]):
        keys = [_key(c) for c in row]
        name_ix = next((j for j, k in enumerate(keys) if k in _NAME_KEYS), None)
        if name_ix is None:
            continue
        mapping = {"name": name_ix}
        for field, aliases in (
            ("price", _PRICE_KEYS),
            ("duration", _DURATION_KEYS),
            ("category", _CATEGORY_KEYS),
            ("code", _CODE_KEYS),
        ):
            ix = next((j for j, k in enumerate(keys) if k in aliases and j != name_ix), None)
            if ix is not None:
                mapping[field] = ix
        return i, mapping
    return -1, {}


# How much of each sheet the layout model sees. Enough to spot a header buried under
# a title block, small enough that a 40-sheet workbook is still one cheap call.
_PREVIEW_ROWS = 15
_PREVIEW_COLS = 12

_LAYOUT_SYSTEM = """You read spreadsheets exported from booking and salon software and
work out where the service menu lives.

You are given a preview of every sheet in one workbook: the first rows of each, with
0-based row and column indexes on every cell.

Find the ONE sheet holding the list of bookable services, and say which column holds
what. Ignore sheets listing locations, staff, customers, appointments, or prose
documentation.

Return ONLY a JSON object, no prose, no markdown fence:
{"sheet": "<exact sheet name>",
 "header_row": <0-based index of the header row, or -1 if there is no header>,
 "columns": {"name": <index>, "price": <index or null>, "duration": <index or null>,
             "category": <index or null>, "code": <index or null>}}

Rules:
- "sheet" must be copied exactly from a given sheet name.
- "name" is required: the column holding what a customer would ask for by name.
  Never point it at a description, notes, or instructions column.
- "duration" is how long the service takes, in minutes. Only pick a column whose
  values are lengths of time. A column of prices is not a duration.
- "price" is what the customer pays. Use null when the file has no prices — never
  substitute another number column just to fill the field.
- Use null for anything not present. Do not invent columns.
- Indexes are 0-based column positions, exactly as shown in the preview."""


def _preview_for_model(all_rows: dict[str, list[list[Any]]]) -> str:
    """Render the workbook as indexed text for the layout model."""
    parts: list[str] = []
    for sname, rows in all_rows.items():
        parts.append(f"### SHEET {sname or '(single sheet)'} — {len(rows)} rows")
        for i, row in enumerate(rows[:_PREVIEW_ROWS]):
            cells = []
            for j, c in enumerate(row[:_PREVIEW_COLS]):
                if c is None:
                    continue
                txt = re.sub(r"\s+", " ", str(c).strip())[:40]
                if txt:
                    cells.append(f"[{j}]={txt}")
            parts.append(f"row {i}: " + (" | ".join(cells) if cells else "(blank)"))
    return "\n".join(parts)


def _llm_locate_layout(
    all_rows: dict[str, list[list[Any]]], model: Optional[str] = None
) -> Optional[dict]:
    """Ask the model where the service list is when the header aliases don't match.

    The keyword pass only recognises headers we thought of — "ServiceName", "Service
    Name", "Service". A salon exporting from something we've never seen may label the
    column "Treatment", "Menu Item" or "Tx", put the data under a merged title block,
    or ship no header at all, and every one of those is a dead end for a fixed alias
    list.

    The model only says WHERE to look. Python still does the reading, so a wrong answer
    produces wrong columns the user can see and correct in the preview, never invented
    services. Returns None on any failure — the caller then reports what it couldn't
    find, exactly as before.
    """
    if not all_rows:
        return None
    try:
        import llm_provider

        reply = llm_provider.chat(
            model=(model or IMPORT_LLM_MODEL),
            messages=[
                {"role": "system", "content": _LAYOUT_SYSTEM},
                {"role": "user", "content": _preview_for_model(all_rows)[:24000]},
            ],
            max_tokens=400,
            temperature=0,
        )
    except Exception as e:
        logger.warning("service_import_layout_failed err=%s", type(e).__name__)
        return None

    obj = _extract_json_object(reply)
    if not obj:
        return None

    # Everything below is validation. The model picks; none of its numbers are trusted.
    wanted = str(obj.get("sheet") or "").strip()
    sheet_name = wanted if wanted in all_rows else None
    if sheet_name is None:
        sheet_name = next((s for s in all_rows if s.lower() == wanted.lower()), None)
    if sheet_name is None and len(all_rows) == 1:
        sheet_name = next(iter(all_rows))
    if sheet_name is None:
        return None

    rows = all_rows[sheet_name]
    width = max((len(r) for r in rows), default=0)
    raw_cols = obj.get("columns")
    if not rows or not width or not isinstance(raw_cols, dict):
        return None

    def _ix(field: str) -> Optional[int]:
        v = raw_cols.get(field)
        # bool is an int subclass, and True would read as column 1.
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            return None
        v = int(v)
        return v if 0 <= v < width else None

    name_ix = _ix("name")
    if name_ix is None:
        return None
    cols: dict[str, int] = {"name": name_ix}
    for field in ("price", "duration", "category", "code"):
        ix = _ix(field)
        if ix is not None and ix != name_ix:
            cols[field] = ix

    hr = obj.get("header_row")
    header_ix = int(hr) if isinstance(hr, (int, float)) and not isinstance(hr, bool) else -1
    # -1 means "data starts at row 0"; extraction reads from header_ix + 1 either way.
    if not -1 <= header_ix < len(rows):
        header_ix = -1
    return {"sheet": sheet_name, "header_row": header_ix, "columns": cols}


_LINK_SYSTEM = """You match salon add-on charges to the services they may be applied to.

You are given a list of SERVICES, a list of ADD-ONS, and NOTES copied from the
business's own service documentation. The notes explain, in prose, which services
each add-on belongs with.

Return ONLY a JSON object, no prose, no markdown fence:
{"links": [{"addon": "<exact add-on name>", "services": ["<exact service name>", ...]}]}

Rules:
- Use names EXACTLY as given in the ADD-ONS and SERVICES lists. Never invent or
  reword a name. If you can't match a name exactly, leave it out.
- Only include an add-on when the notes actually say what it applies to. Omit an
  add-on entirely rather than guessing — an omitted add-on stays available for
  everything, which is the safe default.
- Notes may describe a rule rather than a list, e.g. "all services marked with an
  asterisk". Apply the rule to the SERVICES list yourself and return the matching
  service names.
- An add-on never applies to another add-on.
- If nothing can be determined, return {"links": []}."""


def suggest_addon_links(
    services: list[dict], notes: str, model: Optional[str] = None
) -> dict[str, list[str]]:
    """Ask the model which services each add-on belongs with.

    The keyword pass can tell that "Master Stylist 10" is a charge; only the prose
    says it's for chemical services. Those relationships live in the business's own
    documentation — usually a second sheet in the same workbook — so this reads them
    rather than making an operator tick boxes for a dozen add-ons across every store.

    Returns {addon_name: [service_name, ...]}. Empty on any failure: suggestions are
    a convenience, and no suggestion (add-on available everywhere) is the safe
    fallback. Names are validated against the real lists before being returned, so a
    hallucinated service can't reach the config.
    """
    addons = [s["name"] for s in services if s.get("is_addon") and s.get("name")]
    bookable = [s["name"] for s in services if not s.get("is_addon") and s.get("name")]
    if not addons or not bookable or not (notes or "").strip():
        return {}
    try:
        import llm_provider

        reply = llm_provider.chat(
            model=(model or IMPORT_LLM_MODEL),
            messages=[
                {"role": "system", "content": _LINK_SYSTEM},
                {
                    "role": "user",
                    "content": (
                        "SERVICES:\n" + "\n".join(f"- {n}" for n in bookable)
                        + "\n\nADD-ONS:\n" + "\n".join(f"- {n}" for n in addons)
                        + "\n\nNOTES:\n" + notes[:12000]
                    ),
                },
            ],
            max_tokens=2000,
            temperature=0,
        )
    except Exception as e:
        logger.warning("addon_link_suggest_failed err=%s", type(e).__name__)
        return {}

    obj = _extract_json_object(reply)
    if not obj or not isinstance(obj.get("links"), list):
        return {}
    addon_lookup = {n.lower(): n for n in addons}
    service_lookup = {n.lower(): n for n in bookable}
    out: dict[str, list[str]] = {}
    for link in obj["links"][:100]:
        if not isinstance(link, dict):
            continue
        addon = addon_lookup.get(str(link.get("addon") or "").strip().lower())
        if not addon:
            continue
        matched = []
        for s in link.get("services") or []:
            real = service_lookup.get(str(s or "").strip().lower())
            if real and real not in matched:
                matched.append(real)
        # An empty match list would read as "offer this with nothing" — drop it and
        # let the add-on stay universally available instead.
        if matched:
            out[addon] = matched[:100]
    return out


def _extract_json_object(text: str) -> Optional[dict]:
    """Pull a JSON object out of a model reply, tolerating a code fence or stray prose."""
    import json

    s = (text or "").strip()
    if not s:
        return None
    fence = re.search(r"```(?:json)?\s*(.+?)```", s, re.DOTALL)
    if fence:
        s = fence.group(1).strip()
    try:
        obj = json.loads(s)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass
    start, end = s.find("{"), s.rfind("}")
    if 0 <= start < end:
        try:
            obj = json.loads(s[start : end + 1])
            return obj if isinstance(obj, dict) else None
        except Exception:
            return None
    return None


def _extract_services(
    rows: list[list[Any]], header_ix: int, cols: dict[str, int]
) -> tuple[list[dict], int, bool]:
    """Read service rows given a located header and column map.

    Returns (services, skipped, hit_row_cap). header_ix of -1 means the data starts at
    row 0 — a sheet with no header at all.
    """
    out: list[dict] = []
    skipped = 0
    capped = False
    for row in rows[header_ix + 1 :]:
        if len(out) >= MAX_ROWS:
            capped = True
            break
        if not row:
            continue
        name_ix = cols["name"]
        raw_name = row[name_ix] if name_ix < len(row) else None
        # Guard the None explicitly: str(None) is "None", which is truthy and would
        # sail through as a service literally called "None".
        svc_name = (
            re.sub(r"\s+", " ", str(raw_name).strip())[:200] if raw_name is not None else ""
        )
        if not svc_name or _key(svc_name) in _NAME_KEYS:
            skipped += 1
            continue

        def cell(field: str, _row=row):
            ix = cols.get(field)
            return _row[ix] if ix is not None and ix < len(_row) else None

        duration = max(0, min(_to_int(cell("duration"), 30), 480))
        category = str(cell("category") or "").strip()[:120]
        out.append(
            {
                "name": svc_name,
                "price": round(_to_float(cell("price")), 2),
                # 0 is meaningful here (an add-on charge with no time of its own), so
                # it's preserved; the UI floors it to 5 for anything actually bookable.
                "duration_minutes": duration,
                "category": category,
                "code": str(cell("code") or "").strip()[:64],
                "is_addon": _looks_like_addon(svc_name, duration, category),
            }
        )
    return out, skipped, capped


def _notes_from(all_rows: dict[str, list[list[Any]]], chosen: str) -> str:
    """Every sheet except the service list, flattened to text for the add-on linker."""
    lines: list[str] = []
    for other, rows in all_rows.items():
        if other == chosen:
            continue
        for r in rows:
            cells = [str(c).strip() for c in r if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
            if len(lines) > 400:
                break
    return "\n".join(lines)


def parse_service_spreadsheet(
    data: bytes, filename: str = "", sheet: Optional[str] = None
) -> dict:
    """Parse an .xlsx/.csv service list.

    Returns {"services": [...], "warnings": [...], "sheets": [...], "sheet": str}.
    Never raises for a bad file — it reports the problem so the UI can say what's
    wrong rather than 500ing.
    """
    warnings: list[str] = []
    if not data:
        return {"services": [], "warnings": ["The file was empty."], "sheets": [], "sheet": ""}
    if len(data) > MAX_BYTES:
        return {
            "services": [],
            "warnings": ["That file is too large — export just the service list."],
            "sheets": [],
            "sheet": "",
        }

    all_rows: dict[str, list[list[Any]]] = {}
    sheets: list[str] = []
    chosen = ""
    name = (filename or "").lower()
    try:
        if name.endswith(".csv") or name.endswith(".txt"):
            import csv

            text = data.decode("utf-8-sig", errors="replace")
            all_rows = {"": [list(r) for r in csv.reader(io.StringIO(text))]}
        else:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            sheets = list(wb.sheetnames)
            # Read every sheet up front. The layout fallback below needs to see all of
            # them to choose, and a service list is kilobytes.
            for s in sheets:
                all_rows[s] = [list(r) for r in wb[s].iter_rows(values_only=True)]
    except Exception as e:
        logger.warning("service_import_read_failed file=%r err=%s", filename, e)
        return {
            "services": [],
            "warnings": ["Couldn't read that file. Please upload a .xlsx or .csv export."],
            "sheets": sheets,
            "sheet": chosen,
        }

    # Prefer the requested sheet, else score every sheet and take the richest. A
    # workbook usually has several tabs, and the first one with a name-ish column is
    # often the wrong one — a descriptions tab beats the actual catalogue on a
    # first-match rule. Columns like duration and price identify a real service list.
    if sheet and sheet in all_rows:
        chosen = sheet
    else:
        best_score = -1
        for candidate, candidate_rows in all_rows.items():
            ix, candidate_cols = _find_header(candidate_rows)
            if ix < 0:
                continue
            score = (
                len(candidate_cols)
                + (2 if "duration" in candidate_cols else 0)
                + (2 if "price" in candidate_cols else 0)
            )
            if score > best_score:
                best_score, chosen = score, candidate
        if best_score < 0 and all_rows:
            chosen = next(iter(all_rows))

    rows = all_rows.get(chosen, [])
    header_ix, cols = _find_header(rows)
    out: list[dict] = []
    skipped = 0
    capped = False
    if header_ix >= 0:
        out, skipped, capped = _extract_services(rows, header_ix, cols)

    # Fall back to the model when the fixed aliases got nowhere: an unfamiliar header
    # ("Treatment", "Menu Item"), a sheet with no header at all, or a header we matched
    # that turned out to hold no services. Every export labels these differently and we
    # can't enumerate them all in advance.
    used_ai_layout = False
    if not out:
        layout = _llm_locate_layout(all_rows)
        if layout:
            chosen = layout["sheet"]
            rows = all_rows.get(chosen, [])
            out, skipped, capped = _extract_services(
                rows, layout["header_row"], layout["columns"]
            )
            used_ai_layout = bool(out)

    if not out:
        return {
            "services": [],
            "warnings": [
                "Couldn't find a service list in that file. The sheet needs a column of "
                "service names — a header like 'Service Name' helps."
            ],
            "sheets": sheets,
            "sheet": chosen,
        }

    if capped:
        warnings.append(f"Only the first {MAX_ROWS} services were read.")
    if used_ai_layout:
        warnings.append(
            "This file's layout wasn't one we recognised, so the columns were worked "
            "out from the file itself — check the names and durations below."
        )

    # The service list says what exists; a companion sheet usually says how it's used
    # ("Master Stylist Charge for Chemical Services"), which is where add-on
    # applicability actually lives.
    notes_text = _notes_from(all_rows, chosen)

    # Suggest which services each add-on belongs with, read from the workbook's other
    # sheets. Best-effort: no suggestion just means the add-on stays available for
    # everything, and every one is reviewable before import.
    links = suggest_addon_links(out, notes_text) if notes_text else {}
    for s in out:
        s["applies_to_names"] = links.get(s["name"], []) if s.get("is_addon") else []
    if links:
        warnings.append(
            f"Suggested which services {len(links)} add-on(s) go with, based on the notes "
            "in your file — check them after importing."
        )

    if skipped:
        warnings.append(f"{skipped} row(s) were skipped — no service name.")
    if out and not any(s["price"] for s in out):
        warnings.append(
            "No prices were found in this file — you can add them after importing, or "
            "the receptionist will say the shop confirms pricing when booking."
        )
    return {"services": out, "warnings": warnings, "sheets": sheets, "sheet": chosen}
